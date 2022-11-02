from helpers import settings
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

#function to get values from table reference
def get_table_values(reference, index):
    table = settings.soup.find_all("table", {"ref":reference})[index]
    headings = table.find_all("heading")
    tablerows = table.find_all("tablerow")
    return table,headings,tablerows

#function to write tabledata to spreadsheet
def write_to_sheet(sheet, startingrow, startingcolumn, tabledata, title):
    rownum = startingrow
    sheet.cell(row=rownum, column=startingcolumn, value=title).font = settings.headingfont
    rownum += 1

    for i,heading in enumerate(tabledata[1]):
        sheet.cell(row=rownum, column=i+startingcolumn, value=heading.text).fill = settings.titlefill
    rownum +=1

    for i,row in enumerate(tabledata[2]):
        currow = rownum + i
        tablecells = row.find_all("tablecell")
        for j,cell in enumerate(tablecells):
            sheet.cell(row=currow, column=(j + startingcolumn), value=cell.text)

import copy

#function to create issue sheet
def create_issue_sheet(row, workbook, index):
    items = row.find_all("item")

    title = items[1].text
    print("[!] Generating sheet for issue: " + title)
    title2 = str(index + 1) + " - " + title
    if(len(title2) > 31):
        title2 = title2[:31]
        print("[!] Truncating '" + title + "' to '" + title2 + "'")
    sheet = workbook.create_sheet(title2)

    sheet.cell(row=sheet.max_row + 1, column=1, value="Title").font = settings.headingfont
    sheet.cell(row=sheet.max_row + 1, column=1, value=title)

    section = settings.soup.find("section", {"title": title})

    #remove affected devices section as it was causing issues
    badsections = section.find_all("section", {"title" : "Affected Device"})
    for badsection in badsections:
        badsection.extract()

    findingssection = section.find("section")

    #FINDINGS
    findingssection_copy = copy.copy(findingssection) #create copy to extract from
    findingstables_copy = findingssection_copy.find_all("table")
    for table in findingstables_copy:
         table.extract()
    sheet.cell(row=sheet.max_row + 2, column=1, value="Findings").font = settings.headingfont
    sheet.cell(row=sheet.max_row + 1, column=1, value=findingssection_copy.text.strip())

    #OVERALL
    sheet.cell(row=sheet.max_row + 2, column=1, value="Overall").font = settings.headingfont
    sheet.cell(row=sheet.max_row + 1, column=1, value=section.find("rating").text.strip())

    #DEVICES
    
    sheet.cell(row=sheet.max_row + 2, column=1, value="Affected Devices").font = settings.headingfont
    for device in section.find_all("device"):
        sheet.cell(row=sheet.max_row + 1, column=1, value=device.get("name").strip()) 

    #IMPACT
    sheet.cell(row=sheet.max_row + 2, column=1, value="Impact").font = settings.headingfont
    sheet.cell(row=sheet.max_row + 1, column=1, value=section.find("section", {"ref":"IMPACT"}).text.strip())

    #EASE
    sheet.cell(row=sheet.max_row + 2, column=1, value="Ease").font = settings.headingfont
    sheet.cell(row=sheet.max_row + 1, column=1, value=section.find("section", {"ref":"EASE"}).text.strip())

    #RECOMMENDATION
    sheet.cell(row=sheet.max_row + 2, column=1, value="Recommendation").font = settings.headingfont
    sheet.cell(row=sheet.max_row + 1, column=1, value=section.find("section", {"ref":"RECOMMENDATION"}).text.strip())

    #FINDINGS TABLES
    #check if findings table exists
    findingstables = findingssection.find_all("table")
    #if len(findings) > 0:
    for table in findingstables:
        write_to_sheet(sheet, sheet.max_row + 2, 1, get_table_values(table.get("ref"), 0), "Table - " + table.get("title"))

    return sheet

#fix column widths
def fix_column_width(ws):
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
    ws.column_dimensions = dim_holder
    