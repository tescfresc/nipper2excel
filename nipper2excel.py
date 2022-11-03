from openpyxl import Workbook
import sys
from nipperhelpers import helpers,settings
from argparse import ArgumentParser
from bs4 import BeautifulSoup

#get args
parser = ArgumentParser()
parser.add_argument("-f", "--file", dest="inputfile", help="Input Nipper report XML")
parser.add_argument("-o", "--output", dest="outputfile", help="Output file")
args = parser.parse_args()
def usage():
    print("[*] Nipper2Excel syntax:\n    Run the firewall config through Nipper and save the report (File/Save) as XML, then use\n    python nipper2excel.py -f=<nipper-report-export.xml> -o=<device-ID.xlsx>\n")
if args.inputfile==None or args.outputfile==None:
        print("\n[!] Too few arguments provided.")
        usage()
        sys.exit()

#open input with bs4
ifile = open(args.inputfile, "r", encoding="utf-8")
settings.soup = BeautifulSoup(ifile.read(), features="xml")

#create output workbook
wb = Workbook()
mainws = wb.active
mainws.title = "Main"



#get number of devices, nipper doesn't include a per device issue table or the affected device name on issues if only one device exists
numdevices = len(settings.soup.find_all("table", {"ref":"SECURITY.SUMMARY.AUDITDEVICELIST"}))

#devices table
helpers.write_to_sheet(mainws, mainws.max_row , 1, helpers.get_table_values("SCOPE.AUDITDEVICELIST.TABLE", 0), "Devices")
#risk profile
helpers.write_to_sheet(mainws, mainws.max_row + 2 , 1, helpers.get_table_values("SECURITY.SUMMARY.SECURITYAUDIT.RISKPROFILE", 0), "Risk Profile")
#summary each device
if numdevices > 0:
    helpers.write_to_sheet(mainws, mainws.max_row + 2 , 1, helpers.get_table_values("SECURITY.SUMMARY.AUDITDEVICELIST", 0), "Summary of findings for each device")
#vulnerability audit each device
if len(settings.soup.find_all("table", {"ref":"VULN.SUMMARY.AUDITRESULTLIST"})) > 0:
    helpers.write_to_sheet(mainws, mainws.max_row + 2 , 1, helpers.get_table_values("VULN.SUMMARY.AUDITRESULTLIST", 0), "Summary of findings from the Vulnerability Audit for each device")
    vulnsheet = wb.create_sheet("Vulnerability Audit")
    helpers.write_to_sheet(vulnsheet, 1 , 1, helpers.get_table_values("VULNAUDIT.CONCLUSIONS", 0), "Vulnerability Conclusion")
    for x in range(1, len(settings.soup.find_all("table", {"ref":"VULNAUDIT.CONCLUSIONS"}))):
        helpers.write_to_sheet(vulnsheet, vulnsheet.max_row + 2 , 1, helpers.get_table_values("VULNAUDIT.CONCLUSIONS", x), "Vulnerability List")
    helpers.fix_column_width(vulnsheet)


#create each issue sheet
issues = helpers.get_table_values("SECURITY.FINDINGS.SUMMARY.TABLE", 0)
for i,row in enumerate(issues[2]):
    sheet = helpers.create_issue_sheet(row, wb, i)
    helpers.fix_column_width(sheet)


helpers.fix_column_width(mainws)

wb.save(args.outputfile)




